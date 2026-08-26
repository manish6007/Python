"""Turn what a broker or registrar already gives you into holdings.

Nobody types thirty-five stocks in by hand. Two routes in:

* **Any CSV/XLSX a broker exports.** Rather than making you rename columns to
  match ours, the headers are matched against the names brokers actually use
  and the guessed mapping is shown for correction before anything is
  imported. Renaming still works -- our own names are in the alias list -- it
  just is not required.
* **A CAMS/KFintech consolidated account statement.** One password-protected
  PDF holding every mutual fund folio across both registrars.

Both routes end at the same place: a preview you confirm. Nothing is written
to the portfolio until you have seen the rows.
"""
import csv
import io
import re

# What a column can mean, and the header names brokers really use for it.
# Matched after normalisation (lowercased, punctuation stripped), so
# "Avg. Cost", "avg cost" and "AVG_COST" are the same key.
COLUMN_ALIASES = {
    "identifier": [
        "symbol", "instrument", "tradingsymbol", "trading symbol", "scrip",
        "scrip name", "scrip code", "stock", "stock symbol", "ticker",
        "nse symbol", "nse code", "bse code", "isin", "isin code",
        "security id", "identifier", "folio", "folio no", "folio number",
    ],
    "name": [
        "name", "company", "company name", "security name", "instrument name",
        "stock name", "scheme name", "scheme", "particulars", "description",
    ],
    "units": [
        "qty", "quantity", "shares", "units", "holding qty", "held qty",
        "total qty", "free qty", "balance units", "closing balance",
        "closing unit balance", "no of shares", "quantity available",
    ],
    "avg_cost": [
        "avg", "avg cost", "average price", "avg price", "buy avg",
        "buy average", "average cost", "cost per unit", "avg buy price",
        "purchase nav", "avg nav", "average nav", "buy price", "rate",
    ],
    "last_price": [
        "ltp", "last price", "current price", "market price", "closing price",
        "cmp", "nav", "current nav", "last traded price", "price",
    ],
    "invested": [
        "invested", "invested value", "investment", "investment value",
        "total cost", "cost value", "buy value", "purchase value",
        "amount invested", "total cost value", "purchase cost",
        "value at cost", "inv value", "inv amt", "invested amt",
    ],
    "current_value": [
        "current value", "market value", "present value", "valuation",
        "closing value", "market val", "value",
        # Zerodha writes "Cur. val"; others abbreviate differently again.
        "cur val", "curr val", "current val", "cur value", "curr value",
        "value at market price", "mkt value", "mkt val", "current amt",
    ],
    "purchase_date": [
        "purchase date", "buy date", "date", "transaction date",
        "date of purchase", "trade date",
    ],
}

# Fields the UI offers in the mapping dropdowns, in display order.
MAPPABLE = ["identifier", "name", "units", "avg_cost", "last_price",
            "invested", "current_value", "purchase_date"]


def normalise_header(h):
    h = re.sub(r"[^a-z0-9 ]+", " ", str(h or "").lower())
    return re.sub(r"\s+", " ", h).strip()


def sniff_columns(headers):
    """Guess which of our fields each column holds.

    Exact alias matches win over partial ones, and each field is claimed at
    most once, so a sheet with both "Price" and "Last Price" does not map
    both to last_price.
    """
    norm = [normalise_header(h) for h in headers]
    mapping, taken = {}, set()
    for field in MAPPABLE:                       # exact matches first
        aliases = COLUMN_ALIASES[field]
        for i, h in enumerate(norm):
            if i in taken or not h:
                continue
            if h in aliases:
                mapping[field] = headers[i]
                taken.add(i)
                break
    for field in MAPPABLE:                       # then contains-matches
        if field in mapping:
            continue
        for i, h in enumerate(norm):
            if i in taken or not h:
                continue
            if any(a in h or h in a for a in COLUMN_ALIASES[field]
                   if len(a) > 3):
                mapping[field] = headers[i]
                taken.add(i)
                break
    return mapping


def to_number(value):
    """Parse a spreadsheet money/quantity cell. Returns None when not a number."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in {"-", "--", "NA", "N/A", "nil"}:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s.strip("()"))
    if not s or s in {"-", ".", "-."}:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def read_table(data, filename=""):
    """Rows of a CSV or XLSX as (headers, list-of-dicts)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _read_xlsx(data)
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return [], []
    # Broker exports often carry title/disclaimer lines above the real header.
    header_idx = max(range(min(len(rows), 15)), key=lambda i: len(rows[i]))
    headers = [h.strip() for h in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1:]:
        if len(r) < 2:
            continue
        out.append({headers[i]: (r[i] if i < len(r) else "")
                    for i in range(len(headers))})
    return headers, out


def _read_xlsx(data):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Reading .xlsx needs the openpyxl package; export "
                         "the sheet as CSV instead.")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[("" if c is None else c) for c in row]
            for row in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return [], []
    header_idx = max(range(min(len(rows), 15)), key=lambda i: len(
        [c for c in rows[i] if str(c).strip()]))
    headers = [str(h).strip() for h in rows[header_idx]]
    out = [{headers[i]: (r[i] if i < len(r) else "")
            for i in range(len(headers))} for r in rows[header_idx + 1:]]
    return headers, out


def build_rows(records, mapping, asset_class="stock", owner="Me"):
    """Apply a mapping to raw records and derive what is missing.

    Brokers report either an average price or a total invested value, rarely
    both, so whichever is absent is derived from the other. Rows without a
    usable quantity or name are reported rather than silently dropped.
    """
    out, skipped = [], []
    for i, rec in enumerate(records):
        def val(field):
            col = mapping.get(field)
            return rec.get(col) if col else None

        name = (str(val("name") or "").strip()
                or str(val("identifier") or "").strip())
        units = to_number(val("units"))
        avg = to_number(val("avg_cost"))
        last = to_number(val("last_price"))
        invested = to_number(val("invested"))
        current = to_number(val("current_value"))

        if units and not avg and invested:
            avg = invested / units
        if units and not last and current:
            last = current / units
        if not last:
            last = avg

        if not name:
            skipped.append("row %d: no name or symbol" % (i + 2))
            continue
        if not units or units <= 0:
            skipped.append("row %d (%s): quantity is %s" %
                           (i + 2, name[:30], val("units") or "missing"))
            continue

        ident = str(val("identifier") or "").strip()
        pdate = str(val("purchase_date") or "").strip()
        out.append({
            "owner": owner, "asset_class": asset_class, "name": name[:120],
            "identifier": ident[:60], "units": round(units, 4),
            "avg_cost": round(avg or 0.0, 4), "last_price": round(last or 0.0, 4),
            "invested": round((invested if invested is not None
                               else units * (avg or 0)), 2),
            "current_value": round((current if current is not None
                                    else units * (last or 0)), 2),
            "purchase_date": (pdate[:10]
                              if re.match(r"\d{4}-\d{2}-\d{2}", pdate)
                              else ""),
        })
    return out, skipped


# --------------------------------------------------------------------------
# CAMS / KFintech consolidated account statement
# --------------------------------------------------------------------------
FOLIO_RE = re.compile(r"Folio\s*No[:.\s]*([0-9][\w/\- ]{2,25})", re.I)
CLOSING_RE = re.compile(r"Closing\s*Unit\s*Balance[:.\s]*([\d,]+\.?\d*)", re.I)
NAV_RE = re.compile(r"NAV\s*on\s*[\d\-A-Za-z]+[:.\s]*(?:INR|Rs\.?)?\s*"
                    r"([\d,]+\.?\d*)", re.I)
VALUATION_RE = re.compile(r"Valuation\s*on\s*[\d\-A-Za-z]+[:.\s]*"
                          r"(?:INR|Rs\.?)?\s*([\d,]+\.?\d*)", re.I)
COST_RE = re.compile(r"Total\s*Cost\s*Value[:.\s]*(?:INR|Rs\.?)?\s*"
                     r"([\d,]+\.?\d*)", re.I)
ISIN_RE = re.compile(r"\b(IN[A-Z0-9]{10})\b")
SCHEME_RE = re.compile(r"^\s*([A-Z0-9]{2,10})\s*[-–]\s*(.{6,110}?)\s*"
                       r"(?:\(|Registrar|ISIN|$)", re.M)


def extract_cas_text(pdf_bytes, password=""):
    """Text of a CAS PDF, opening it with the password if it has one."""
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ValueError("Reading a CAS needs the pypdf package "
                         "(pip install -r requirements.txt).")
    reader = PdfReader(io.BytesIO(pdf_bytes))
    if reader.is_encrypted:
        if not reader.decrypt(password or ""):
            raise PermissionError(
                "Wrong password. A CAS is usually locked with the password "
                "you chose when requesting it from CAMS or KFintech.")
    return "\n".join((p.extract_text() or "") for p in reader.pages)


def parse_cas(text, owner="Me"):
    """Holdings from CAS text.

    Statements differ between registrars and change format, so this reports
    what it could not read instead of guessing. Every row is shown for
    confirmation before anything is imported.
    """
    rows, notes = [], []
    blocks = re.split(r"(?=Folio\s*No)", text, flags=re.I)
    for block in blocks:
        folio_m = FOLIO_RE.search(block)
        if not folio_m:
            continue
        units_m = CLOSING_RE.search(block)
        if not units_m:
            continue
        units = to_number(units_m.group(1))
        if not units or units <= 0:
            continue                      # closed folio, nothing held
        scheme_m = SCHEME_RE.search(block)
        name = (scheme_m.group(2).strip() if scheme_m
                else "Scheme (name not read)")
        nav = to_number(NAV_RE.search(block).group(1)) if NAV_RE.search(block) else None
        val = (to_number(VALUATION_RE.search(block).group(1))
               if VALUATION_RE.search(block) else None)
        cost = (to_number(COST_RE.search(block).group(1))
                if COST_RE.search(block) else None)
        isin = ISIN_RE.search(block)
        if not nav and val and units:
            nav = val / units
        if not name or name.startswith("Scheme ("):
            notes.append("A folio's scheme name could not be read — set it "
                         "by hand after importing.")
        rows.append({
            "owner": owner, "asset_class": "mutual_fund",
            "name": name[:120],
            "identifier": folio_m.group(1).strip()[:60],
            "isin": isin.group(1) if isin else "",
            "units": round(units, 4),
            "avg_cost": round(cost / units, 4) if cost and units else 0.0,
            "last_price": round(nav or 0.0, 4),
            "invested": round(cost or 0.0, 2),
            "current_value": round(val or (units * (nav or 0)), 2),
            "purchase_date": "",
        })
    if not rows:
        notes.append("No folios with a closing balance were found. If this is "
                     "a summary-only statement, request the detailed one; "
                     "otherwise the format may have changed — the CSV route "
                     "still works.")
    return rows, sorted(set(notes))

# --------------------------------------------------------------------------
# CAS "Consolidated Account Summary" -- the table format
#
# Columns: Folio No. | ISIN | Scheme Name | Cost Value | Unit Balance |
#          NAV Date | NAV | Market Value | Registrar
#
# Extracted text wraps scheme names over several lines and often runs the
# folio straight into the ISIN, so rows are anchored on the ISIN -- exactly
# one per holding, and unmistakable -- rather than on line breaks. Every
# money column carries decimals while the digits inside scheme names
# ("NASDAQ 100", "Nifty 50", the registrar's "128TSDGG" prefix) do not, which
# is what makes the four numbers at the end of a row safe to read positionally.
# --------------------------------------------------------------------------


# Indian ISINs are IN + E/F/9 + 9 more. A word boundary cannot be used at
# the start: extracted text frequently glues the folio to the ISIN with no
# space ("90722941761/0INF846K01EW2"), which silently dropped those rows.
ISIN_TOKEN = re.compile(r"(?<![A-Z])IN[EF0-9][A-Z0-9]{9}(?![A-Z0-9])")
DECIMAL_TOKEN = re.compile(r"\d[\d,]*\.\d+")
NAV_DATE_TOKEN = re.compile(r"\d{2}-[A-Za-z]{3}-\d{4}")
SCHEME_CODE_PREFIX = re.compile(r"^[A-Z0-9]{2,12}\s*[-\u2013]\s*")
REGISTRAR_TOKEN = re.compile(r"\b(CAMS|KFINTECH|KARVY)\b", re.I)


def _clean_scheme(text):
    text = re.sub(r"\s+", " ", text).strip(" -\u2013:")
    text = SCHEME_CODE_PREFIX.sub("", text)          # drop "128TSDGG - "
    text = re.sub(r"\((?:Non[\s-]?Demat|Demat)\)", "", text, flags=re.I)
    return re.sub(r"\s+", " ", text).strip(" -\u2013:,")


def parse_cas_summary(text, owner="Me"):
    """Holdings from a Consolidated Account Summary.

    Returns (rows, notes). Anything that cannot be read is reported rather
    than guessed at, and the caller confirms every row before import.
    """
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    anchors = [i for i, ln in enumerate(lines) if ISIN_TOKEN.search(ln)]
    rows, notes = [], []
    for n, start in enumerate(anchors):
        end = anchors[n + 1] if n + 1 < len(anchors) else len(lines)
        block_text = " ".join(lines[start:end])
        isin_m = ISIN_TOKEN.search(block_text)
        isin = isin_m.group(0)
        before, after = block_text[:isin_m.start()], block_text[isin_m.end():]

        # Folio is whatever sits left of the ISIN, often glued to it.
        folio = re.sub(r"\s+", "", before).strip()
        folio = re.sub(r"^(?:Folio\s*No\.?:?)", "", folio, flags=re.I).strip()

        nums = DECIMAL_TOKEN.findall(after)
        if len(nums) < 4:
            notes.append("A row for ISIN %s had %d numeric columns, not the "
                         "expected four — check it in the preview."
                         % (isin, len(nums)))
            continue
        cost, units, nav, market = (to_number(x) for x in nums[:4])
        if not units or units <= 0:
            continue                       # exited scheme, nothing held

        scheme = _clean_scheme(after[:after.find(nums[0])])
        date_m = NAV_DATE_TOKEN.search(after)
        registrar_m = REGISTRAR_TOKEN.search(after)
        if not scheme:
            scheme = "Scheme (name not read)"
            notes.append("A scheme name could not be read for ISIN %s — set "
                         "it by hand after importing." % isin)
        rows.append({
            "owner": owner, "asset_class": "mutual_fund",
            "name": scheme[:120], "identifier": folio[:60], "isin": isin,
            "units": round(units, 4),
            "avg_cost": round(cost / units, 4) if cost and units else 0.0,
            "last_price": round(nav or 0.0, 4),
            "invested": round(cost or 0.0, 2),
            "current_value": round(market or (units * (nav or 0)), 2),
            "nav_date": date_m.group(0) if date_m else "",
            "registrar": registrar_m.group(1).upper() if registrar_m else "",
            "purchase_date": "",
        })
    return rows, sorted(set(notes))


TOTAL_ROW = re.compile(r"Total\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.\d+)", re.I)


def check_against_total(text, rows):
    """Compare what was parsed against the statement's own Total row.

    A parser that quietly drops rows is worse than one that fails loudly, and
    the summary prints its own totals, so there is no excuse for not checking.
    """
    m = TOTAL_ROW.search(text or "")
    if not m:
        return []
    stated_cost, stated_market = to_number(m.group(1)), to_number(m.group(2))
    got_cost = sum(r["invested"] for r in rows)
    got_market = sum(r["current_value"] for r in rows)
    notes = []
    for label, stated, got in (("cost", stated_cost, got_cost),
                               ("market value", stated_market, got_market)):
        if stated and abs(stated - got) > max(1.0, stated * 0.005):
            notes.append(
                "The statement's total %s is %s but the rows read add up to "
                "%s — a difference of %s. Some holdings were not read; check "
                "the preview against your statement before importing."
                % (label, _fmt(stated), _fmt(got), _fmt(abs(stated - got))))
    return notes


def _fmt(x):
    return "{:,.2f}".format(x or 0)


def parse_cas_any(text, owner="Me"):
    """Try the summary table first, then the detailed statement layout.

    CAMS and KFintech issue both; people request whichever they find, so the
    importer should not care which arrived.
    """
    rows, notes = parse_cas_summary(text, owner=owner)
    if rows:
        return rows, sorted(set(notes + check_against_total(text, rows))), \
            "summary"
    rows, notes = parse_cas(text, owner=owner)
    if rows:
        return rows, notes, "detailed"
    return [], ["No holdings could be read. This importer understands the "
                "CAMS/KFintech Consolidated Account Summary table and the "
                "detailed statement; if yours looks different, the broker "
                "CSV route still works."], "unknown"
